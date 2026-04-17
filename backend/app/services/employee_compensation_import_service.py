from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook

from ..models.models import Employee


DEFAULT_COMPENSATION_MODE = "hourly"
DEFAULT_PERDIEM_THRESHOLD_HOURS = 7.0
DEFAULT_PAYROLL_COMPANY = "254981"


@dataclass
class EmployeeCompensationRow:
    name: str
    hourly_rate: float | None
    perdiem: float | None
    compensation_mode: str
    perdiem_mode: str
    perdiem_threshold_hours: float
    raw_mode: str = ""


def _clean_str(value) -> str:
    return str(value or "").strip()


def _normalize_name(value: str) -> str:
    raw = unicodedata.normalize("NFKD", _clean_str(value).lower())
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return re.sub(r"\s+", " ", raw).strip()


def _name_variants(value: str) -> list[str]:
    raw = _clean_str(value)
    normalized = _normalize_name(raw)
    variants: list[str] = []
    if normalized:
        variants.append(normalized)
        sorted_tokens = " ".join(sorted(normalized.split()))
        if sorted_tokens and sorted_tokens not in variants:
            variants.append(sorted_tokens)

    if "," in raw:
        last_name, first_name = [part.strip() for part in raw.split(",", 1)]
        swapped = _normalize_name(f"{first_name} {last_name}")
        if swapped and swapped not in variants:
            variants.append(swapped)
        sorted_swapped = " ".join(sorted(swapped.split()))
        if sorted_swapped and sorted_swapped not in variants:
            variants.append(sorted_swapped)

    return variants


def _parse_number(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    cleaned = str(value).strip().replace(",", ".")
    if not cleaned:
        return None
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return None


def _normalize_headers(raw_row: dict[str, object]) -> dict[str, object]:
    return {_normalize_name(key): value for key, value in raw_row.items()}


def _extract_perdiem_threshold(raw_mode: str) -> float:
    matches = re.findall(r"(\d+(?:[.,]\d+)?)", raw_mode or "")
    if not matches:
        return DEFAULT_PERDIEM_THRESHOLD_HOURS
    try:
        return max(0, float(matches[0].replace(",", ".")))
    except ValueError:
        return DEFAULT_PERDIEM_THRESHOLD_HOURS


def _parse_compensation_modes(raw_mode: str, perdiem_value: float | None) -> tuple[str, str, float]:
    mode = _normalize_name(raw_mode)
    if "honoraire" in mode:
        return ("honoraires", "", 0.0)
    if "perdiem horaire" in mode or ("horaire" in mode and "perdiem" in mode):
        return (DEFAULT_COMPENSATION_MODE, "hourly", 0.0)
    if "quart" in mode or "jour" in mode:
        return (
            DEFAULT_COMPENSATION_MODE,
            "per_shift_min_hours",
            _extract_perdiem_threshold(raw_mode),
        )
    if (perdiem_value or 0) > 0:
        return (
            DEFAULT_COMPENSATION_MODE,
            "per_shift_min_hours",
            DEFAULT_PERDIEM_THRESHOLD_HOURS,
        )
    return (DEFAULT_COMPENSATION_MODE, "", DEFAULT_PERDIEM_THRESHOLD_HOURS)


def _iter_rows_from_xlsx(file_bytes: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean_str(cell) for cell in rows[0]]
    items: list[dict[str, object]] = []
    for raw_row in rows[1:]:
        row = {
            headers[index]: raw_row[index]
            for index in range(min(len(headers), len(raw_row)))
            if headers[index]
        }
        if any(_clean_str(value) for value in row.values()):
            items.append(row)
    return items


def _iter_rows_from_csv(file_bytes: bytes) -> list[dict[str, object]]:
    text = file_bytes.decode("utf-8-sig")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [row for row in reader if any(_clean_str(value) for value in row.values())]


def parse_employee_compensation_file(file_bytes: bytes, filename: str) -> list[EmployeeCompensationRow]:
    extension = _clean_str(filename).lower().rsplit(".", 1)[-1] if "." in _clean_str(filename) else ""
    if extension in {"xlsx", "xlsm", "xltx", "xltm"}:
        raw_rows = _iter_rows_from_xlsx(file_bytes)
    elif extension in {"csv", "txt"}:
        raw_rows = _iter_rows_from_csv(file_bytes)
    else:
        raise ValueError("Format non supporte. Utilise un fichier .xlsx ou .csv.")

    parsed_rows: list[EmployeeCompensationRow] = []
    for raw_row in raw_rows:
        normalized = _normalize_headers(raw_row)
        name = _clean_str(
            normalized.get("nom de la ressource")
            or normalized.get("nom")
            or normalized.get("ressource")
        )
        hourly_rate = _parse_number(normalized.get("taux horaire"))
        perdiem = _parse_number(normalized.get("perdiem"))
        raw_mode = _clean_str(normalized.get("mode"))
        if not name:
            continue
        if hourly_rate is None and perdiem is None and not raw_mode:
            continue
        compensation_mode, perdiem_mode, threshold = _parse_compensation_modes(raw_mode, perdiem)
        parsed_rows.append(
            EmployeeCompensationRow(
                name=name,
                hourly_rate=hourly_rate,
                perdiem=perdiem,
                compensation_mode=compensation_mode,
                perdiem_mode=perdiem_mode,
                perdiem_threshold_hours=threshold,
                raw_mode=raw_mode,
            )
        )
    return parsed_rows


def apply_employee_compensation_rows(
    employees: Iterable[Employee],
    rows: Iterable[EmployeeCompensationRow],
) -> dict[str, object]:
    employee_list = list(employees)
    row_list = list(rows)

    lookup: dict[str, list[Employee]] = {}
    for employee in employee_list:
        for variant in _name_variants(getattr(employee, "name", "")):
            lookup.setdefault(variant, []).append(employee)

    matched = 0
    updated = 0
    unmatched: list[dict[str, object]] = []
    ambiguous: list[dict[str, object]] = []
    touched_employee_ids: set[int] = set()

    for row in row_list:
        candidates: list[Employee] = []
        seen_employee_ids: set[int] = set()
        for variant in _name_variants(row.name):
            for employee in lookup.get(variant, []):
                if employee.id in seen_employee_ids:
                    continue
                seen_employee_ids.add(employee.id)
                candidates.append(employee)

        if not candidates:
            unmatched.append(
                {
                    "name": row.name,
                    "hourly_rate": row.hourly_rate,
                    "perdiem": row.perdiem,
                    "mode": row.raw_mode,
                }
            )
            continue
        if len(candidates) > 1:
            ambiguous.append(
                {
                    "name": row.name,
                    "hourly_rate": row.hourly_rate,
                    "perdiem": row.perdiem,
                    "mode": row.raw_mode,
                    "employee_names": ", ".join(employee.name for employee in candidates),
                }
            )
            continue

        employee = candidates[0]
        matched += 1
        touched_employee_ids.add(employee.id)
        changed = False

        normalized_rate = 0.0 if row.compensation_mode == "honoraires" else float(row.hourly_rate or 0)
        normalized_perdiem = 0.0 if row.compensation_mode == "honoraires" else float(row.perdiem or 0)

        if row.hourly_rate is not None or row.compensation_mode == "honoraires":
            if float(getattr(employee, "rate", 0) or 0) != normalized_rate:
                employee.rate = normalized_rate
                changed = True
            if float(getattr(employee, "salary", 0) or 0) != normalized_rate:
                employee.salary = normalized_rate
                changed = True

        if row.perdiem is not None or row.compensation_mode == "honoraires":
            if float(getattr(employee, "perdiem", 0) or 0) != normalized_perdiem:
                employee.perdiem = normalized_perdiem
                changed = True

        if _clean_str(getattr(employee, "payroll_compensation_mode", "")) != row.compensation_mode:
            employee.payroll_compensation_mode = row.compensation_mode
            changed = True
        if _clean_str(getattr(employee, "perdiem_mode", "")) != row.perdiem_mode:
            employee.perdiem_mode = row.perdiem_mode
            changed = True
        if float(getattr(employee, "perdiem_threshold_hours", 0) or 0) != float(row.perdiem_threshold_hours or 0):
            employee.perdiem_threshold_hours = float(row.perdiem_threshold_hours or 0)
            changed = True
        if not _clean_str(getattr(employee, "payroll_company", "")):
            employee.payroll_company = DEFAULT_PAYROLL_COMPANY
            changed = True

        if changed:
            updated += 1

    return {
        "total_rows": len(row_list),
        "matched_rows": matched,
        "updated_employees": updated,
        "touched_employee_ids": sorted(touched_employee_ids),
        "unmatched_rows": unmatched,
        "ambiguous_rows": ambiguous,
    }
