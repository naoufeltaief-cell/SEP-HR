from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook

from ..models.models import Employee

DEFAULT_DESJARDINS_COMPANY = "254981"


@dataclass
class DesjardinsEmployeeRow:
    division: str
    matricule: str
    name: str
    status: str


def _clean_str(value) -> str:
    return str(value or "").strip()


def _normalize_name(value: str) -> str:
    raw = unicodedata.normalize("NFKD", _clean_str(value).lower())
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return re.sub(r"\s+", " ", raw).strip()


def _name_variants(value: str) -> list[str]:
    raw = _clean_str(value)
    normalized = _normalize_name(raw)
    variants: list[str] = []
    if normalized:
        variants.append(normalized)
        sorted_tokens = " ".join(sorted(normalized.split()))
        if sorted_tokens and sorted_tokens not in variants:
            variants.append(sorted_tokens)

    if "," in raw:
        last_name, first_name = [part.strip() for part in raw.split(",", 1)]
        swapped = _normalize_name(f"{first_name} {last_name}")
        if swapped and swapped not in variants:
            variants.append(swapped)
        sorted_swapped = " ".join(sorted(swapped.split()))
        if sorted_swapped and sorted_swapped not in variants:
            variants.append(sorted_swapped)

    return variants


def _iter_rows_from_xlsx(file_bytes: bytes) -> Iterable[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean_str(cell) for cell in rows[0]]
    items: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        row = {
            headers[index]: raw_row[index]
            for index in range(min(len(headers), len(raw_row)))
            if headers[index]
        }
        if any(_clean_str(value) for value in row.values()):
            items.append(row)
    return items


def _iter_rows_from_csv(file_bytes: bytes) -> Iterable[dict[str, str]]:
    text = file_bytes.decode("utf-8-sig")
    sample = text[:2048]
    delimiter = ";"
    if sample.count(",") > sample.count(";"):
        delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [row for row in reader if any(_clean_str(value) for value in row.values())]


def parse_desjardins_employee_file(file_bytes: bytes, filename: str) -> list[DesjardinsEmployeeRow]:
    extension = _clean_str(filename).lower().rsplit(".", 1)[-1] if "." in _clean_str(filename) else ""
    if extension in {"xlsx", "xlsm", "xltx", "xltm"}:
        raw_rows = _iter_rows_from_xlsx(file_bytes)
    elif extension in {"csv", "txt"}:
        raw_rows = _iter_rows_from_csv(file_bytes)
    else:
        raise ValueError("Format non supporte. Utilise un fichier .xlsx ou .csv.")

    parsed_rows: list[DesjardinsEmployeeRow] = []
    for raw_row in raw_rows:
        normalized_keys = {_normalize_name(key): value for key, value in raw_row.items()}
        division = _clean_str(normalized_keys.get("division"))
        matricule = _clean_str(normalized_keys.get("matricule"))
        name = _clean_str(normalized_keys.get("nom"))
        status = _clean_str(normalized_keys.get("statut"))
        if not matricule or not name:
            continue
        parsed_rows.append(
            DesjardinsEmployeeRow(
                division=division,
                matricule=matricule,
                name=name,
                status=status,
            )
        )
    return parsed_rows


def apply_desjardins_employee_rows(
    employees: Iterable[Employee],
    rows: Iterable[DesjardinsEmployeeRow],
) -> dict[str, object]:
    employee_list = list(employees)
    row_list = list(rows)

    lookup: dict[str, list[Employee]] = {}
    for employee in employee_list:
        for variant in _name_variants(getattr(employee, "name", "")):
            lookup.setdefault(variant, []).append(employee)

    matched = 0
    updated = 0
    unmatched: list[dict[str, str]] = []
    ambiguous: list[dict[str, str]] = []
    touched_employee_ids: set[int] = set()

    for row in row_list:
        candidates: list[Employee] = []
        seen_employee_ids: set[int] = set()
        for variant in _name_variants(row.name):
            for employee in lookup.get(variant, []):
                if employee.id in seen_employee_ids:
                    continue
                seen_employee_ids.add(employee.id)
                candidates.append(employee)

        if not candidates:
            unmatched.append({"name": row.name, "matricule": row.matricule, "division": row.division})
            continue
        if len(candidates) > 1:
            ambiguous.append(
                {
                    "name": row.name,
                    "matricule": row.matricule,
                    "division": row.division,
                    "employee_names": ", ".join(employee.name for employee in candidates),
                }
            )
            continue

        employee = candidates[0]
        matched += 1
        touched_employee_ids.add(employee.id)
        changed = False
        if _clean_str(getattr(employee, "matricule", "")) != row.matricule:
            employee.matricule = row.matricule
            changed = True
        if _clean_str(getattr(employee, "payroll_division", "")) != row.division:
            employee.payroll_division = row.division
            changed = True
        if _clean_str(getattr(employee, "payroll_company", "")) != DEFAULT_DESJARDINS_COMPANY:
            employee.payroll_company = DEFAULT_DESJARDINS_COMPANY
            changed = True
        if changed:
            updated += 1

    return {
        "total_rows": len(row_list),
        "matched_rows": matched,
        "updated_employees": updated,
        "touched_employee_ids": sorted(touched_employee_ids),
        "unmatched_rows": unmatched,
        "ambiguous_rows": ambiguous,
        "default_company": DEFAULT_DESJARDINS_COMPANY,
    }
