from __future__ import annotations

import io
import os
import sys
import unittest
from types import SimpleNamespace

from openpyxl import Workbook

ROOT = os.path.dirname(__file__)
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from app.services.employee_payroll_import_service import (  # noqa: E402
    DEFAULT_DESJARDINS_COMPANY,
    apply_desjardins_employee_rows,
    parse_desjardins_employee_file,
)


class EmployeePayrollImportTests(unittest.TestCase):
    def test_parse_xlsx_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Division", "Matricule", "Nom", "Statut"])
        sheet.append(["99-1-52-999", 150, "SOUVERAIN, NEPHTALY", "Actif"])
        buffer = io.BytesIO()
        workbook.save(buffer)

        rows = parse_desjardins_employee_file(buffer.getvalue(), "liste.xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].division, "99-1-52-999")
        self.assertEqual(rows[0].matricule, "150")
        self.assertEqual(rows[0].name, "SOUVERAIN, NEPHTALY")

    def test_parse_xlsx_rows_across_multiple_sheets_with_offset_headers(self):
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "Sheet1"
        sheet1.append(["Division", "Matricule", "Nom", "Statut"])
        sheet1.append(["99-1-52-999", 150, "SOUVERAIN, NEPHTALY", "Actif"])

        sheet2 = workbook.create_sheet("Sheet2")
        sheet2.append(["", "", "", ""])
        sheet2.append(["Division", "Matricule", "Nom", "Statut"])
        sheet2.append(["99-1-52-111", 151, "TAIEF, NAOUFEL", "Actif"])

        sheet3 = workbook.create_sheet("Sheet3")
        sheet3.append(["Liste administrative", "", "", ""])
        sheet3.append(["", "", "", ""])
        sheet3.append(["Division", "Matricule", "Nom", "Statut"])
        sheet3.append(["99-1-52-222", 152, "BOUTAINA, TAIEF", "Actif"])

        buffer = io.BytesIO()
        workbook.save(buffer)

        rows = parse_desjardins_employee_file(buffer.getvalue(), "liste.xlsx")

        self.assertEqual(len(rows), 3)
        self.assertEqual([row.matricule for row in rows], ["150", "151", "152"])
        self.assertEqual(rows[1].name, "TAIEF, NAOUFEL")
        self.assertEqual(rows[2].division, "99-1-52-222")

    def test_apply_rows_updates_employee_matricule_and_division(self):
        employees = [
            SimpleNamespace(
                id=1,
                name="Nephtaly Souverain",
                matricule="",
                payroll_company="",
                payroll_division="",
            ),
            SimpleNamespace(
                id=2,
                name="Boutaina Taief",
                matricule="A001",
                payroll_company="254981",
                payroll_division="OLD",
            ),
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Division", "Matricule", "Nom", "Statut"])
        sheet.append(["99-1-52-999", 150, "SOUVERAIN, NEPHTALY", "Actif"])
        sheet.append(["99-999-999-999", 34, "Taief, Boutaina", "Actif"])
        buffer = io.BytesIO()
        workbook.save(buffer)

        rows = parse_desjardins_employee_file(buffer.getvalue(), "liste.xlsx")
        report = apply_desjardins_employee_rows(employees, rows)

        self.assertEqual(report["matched_rows"], 2)
        self.assertEqual(report["updated_employees"], 2)
        self.assertEqual(employees[0].matricule, "150")
        self.assertEqual(employees[0].payroll_division, "99-1-52-999")
        self.assertEqual(employees[0].payroll_company, DEFAULT_DESJARDINS_COMPANY)
        self.assertEqual(employees[1].matricule, "34")
        self.assertEqual(employees[1].payroll_division, "99-999-999-999")


if __name__ == "__main__":
    unittest.main(verbosity=2)
