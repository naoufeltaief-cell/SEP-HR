"""
Targeted regression checks for the Desjardins payroll export.

Run with:
  python test_payroll_export.py
"""

from __future__ import annotations

import io
import os
import sys
import unittest

from openpyxl import load_workbook

ROOT = os.path.dirname(__file__)
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from app.services.payroll_service import (  # noqa: E402
    ALLOWED_PAYROLL_CODES,
    CSV_HEADERS,
    DEFAULT_PAYROLL_COMPANY,
    DEFAULT_PAYROLL_STATEMENT_NUMBER,
    DEFAULT_PAYROLL_TRANSACTION_TYPE,
    _rows_to_csv_bytes,
    _rows_to_xlsx_bytes,
    _week_number_for_date,
    build_desjardins_export_row,
)


class PayrollExportTests(unittest.TestCase):
    def test_build_row_applies_system_defaults_before_validation(self):
        from datetime import date
        from types import SimpleNamespace

        employee = SimpleNamespace(
            id=12,
            matricule="A001",
            payroll_company="",
            payroll_statement_number="",
            payroll_transaction_type="",
            payroll_division="DIV",
            payroll_service="SRV",
            payroll_department="DEP",
            payroll_subdepartment="SUB",
        )

        row = build_desjardins_export_row(
            employee=employee,
            company_context="",
            code="1",
            week_number=1,
            transaction_date=date(2026, 4, 18),
            source_type="schedule",
            source_id="abc",
            export_key="schedule:abc:1",
            sort_order=10,
            quantity=7.5,
        )

        self.assertEqual(row.company, DEFAULT_PAYROLL_COMPANY)
        self.assertEqual(row.statement_number, DEFAULT_PAYROLL_STATEMENT_NUMBER)
        self.assertEqual(row.transaction_type, DEFAULT_PAYROLL_TRANSACTION_TYPE)
        self.assertEqual(row.matricule, "A001")

    def test_build_row_requires_employee_matricule(self):
        from datetime import date
        from types import SimpleNamespace

        employee = SimpleNamespace(
            id=12,
            matricule="",
            payroll_company="",
            payroll_statement_number="",
            payroll_transaction_type="",
            payroll_division="",
            payroll_service="",
            payroll_department="",
            payroll_subdepartment="",
        )

        with self.assertRaisesRegex(ValueError, "matricule est absent du profil employe"):
            build_desjardins_export_row(
                employee=employee,
                company_context="",
                code="1",
                week_number=1,
                transaction_date=date(2026, 4, 18),
                source_type="schedule",
                source_id="abc",
                export_key="schedule:abc:1",
                sort_order=10,
                quantity=7.5,
            )

    def test_allowed_codes_are_strictly_limited(self):
        self.assertEqual(ALLOWED_PAYROLL_CODES, {"1", "4", "43", "57", "517", "518"})

    def test_csv_headers_match_expected_desjardins_layout(self):
        self.assertEqual(
            CSV_HEADERS,
            [
                "Compagnie",
                "Matricule",
                "No relevé",
                "Type transaction",
                "Code gain/déduction",
                "Quantité",
                "Taux",
                "Montant",
                "Semaine",
                "Division",
                "Service",
                "Département",
                "Sous-département",
                "Date de transaction",
            ],
        )

    def test_week_number_split_is_biweekly(self):
        from datetime import date

        period_start = date(2026, 4, 5)
        self.assertEqual(_week_number_for_date(period_start, date(2026, 4, 5)), 1)
        self.assertEqual(_week_number_for_date(period_start, date(2026, 4, 11)), 1)
        self.assertEqual(_week_number_for_date(period_start, date(2026, 4, 12)), 2)
        self.assertEqual(_week_number_for_date(period_start, date(2026, 4, 18)), 2)

    def test_csv_export_uses_semicolon_utf8_and_decimal_dot(self):
        rows = [
            {
                "company": "SEP",
                "matricule": "A001",
                "statement_number": "R1",
                "transaction_type": "G",
                "payroll_code": "57",
                "quantity": 345.5,
                "rate": 0.525,
                "amount": None,
                "week_number": 1,
                "division": "DIV",
                "service": "SRV",
                "department": "DEP",
                "subdepartment": "SUB",
                "transaction_date": "2026-04-18",
            }
        ]
        payload = _rows_to_csv_bytes(rows).decode("utf-8")
        first_line, second_line = payload.strip().split("\n")
        self.assertEqual(first_line.count(";"), 13)
        self.assertIn("Code gain/déduction", first_line)
        self.assertIn("345.5", second_line)
        self.assertIn("0.525", second_line)
        self.assertNotIn(",", second_line)

    def test_xlsx_export_keeps_exact_column_order(self):
        rows = [
            {
                "company": "SEP",
                "matricule": "A001",
                "statement_number": "0",
                "transaction_type": "G",
                "payroll_code": "1",
                "quantity": 37.5,
                "rate": None,
                "amount": None,
                "week_number": 2,
                "division": "DIV",
                "service": "SRV",
                "department": "DEP",
                "subdepartment": "SUB",
                "transaction_date": "2026-04-18",
            }
        ]
        raw = _rows_to_xlsx_bytes(rows)
        workbook = load_workbook(io.BytesIO(raw))
        sheet = workbook.active
        header = [sheet.cell(row=1, column=index).value for index in range(1, 15)]
        data_row = [sheet.cell(row=2, column=index).value for index in range(1, 15)]
        self.assertEqual(header, CSV_HEADERS)
        self.assertEqual(data_row[0], "SEP")
        self.assertEqual(data_row[1], "A001")
        self.assertEqual(data_row[2], "0")
        self.assertEqual(data_row[3], "G")
        self.assertEqual(data_row[4], "1")
        self.assertEqual(data_row[5], 37.5)
        self.assertEqual(data_row[8], 2)
        self.assertEqual(data_row[13], "2026-04-18")


if __name__ == "__main__":
    unittest.main(verbosity=2)
